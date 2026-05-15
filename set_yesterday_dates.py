"""
FIX 1: Pre-run script called by the scheduled task before core.py.
Sets From Date and To Date in control.xlsx to yesterday's date,
resets Download=TRUE for all reports, and re-enables all units
so that every scheduled run always downloads fresh yesterday data.
"""
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

CONTROL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "control.xlsx")

def set_yesterday_dates():
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")
    print(f"[SCHEDULE] Setting run dates to yesterday: {yesterday}")

    wb = load_workbook(CONTROL_PATH)

    # --- Update Main sheet key/value section ---
    ws_main = wb["Main"]
    for row in ws_main.iter_rows():
        cell_key = str(row[0].value or "").strip().lower()
        if cell_key == "from date":
            row[1].value = yesterday
        elif cell_key == "to date":
            row[1].value = yesterday

    # --- Update Report sheet: set all Download = TRUE and both date columns ---
    if "Report" in wb.sheetnames:
        ws_report = wb["Report"]
        headers = [str(cell.value or "").strip() for cell in ws_report[1]]
        try:
            dl_col   = headers.index("Download") + 1
            from_col = headers.index("From Date") + 1 if "From Date" in headers else None
            to_col   = headers.index("To Date")   + 1 if "To Date"   in headers else None
        except ValueError:
            dl_col = from_col = to_col = None

        if dl_col:
            for row in ws_report.iter_rows(min_row=2):
                row[dl_col - 1].value = "TRUE"
                if from_col:
                    row[from_col - 1].value = yesterday
                if to_col:
                    row[to_col - 1].value = yesterday

    wb.save(CONTROL_PATH)
    print(f"[SCHEDULE] control.xlsx updated — From/To Date={yesterday}, Download=TRUE for all reports.")

if __name__ == "__main__":
    set_yesterday_dates()
