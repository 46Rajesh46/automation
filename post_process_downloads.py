import os
import glob
import shutil
import time
import pandas as pd
import pythoncom
import win32com.client as win32
from openpyxl import load_workbook
import re
from datetime import datetime

# === Config ===
control_path = "control.xlsx"
min_rows_required = 5
status_col = "Status"
download_col = "Download"
error_col = "Failure Reason"

CLINIC_ONLY_REPORT_NAMES = {
    'revenue_service_level',
    'sales_day_book_excel',
    'opd_foot_falls_new',
}


def append_run_log(message):
    log_path = os.path.join(os.getcwd(), f"run_log_{datetime.now().strftime('%Y-%m-%d')}.txt")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {message}\n")

# === Load Main sheet ===
config_df = pd.read_excel(control_path, sheet_name="Main", header=None, index_col=0)
# FIX 4: Fall back to dynamic path if control.xlsx has no output folder or still has old D:\ path
_raw_output_folder = str(config_df.loc["Output folder", 1]).strip()
if not _raw_output_folder or _raw_output_folder.lower() == "nan" or _raw_output_folder.startswith("D:\\"):
    output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "download")
else:
    output_folder = _raw_output_folder

# Dashboard folder
dashboard_folder = os.path.join(os.path.dirname(output_folder), "dashboard")
os.makedirs(dashboard_folder, exist_ok=True)

# Read Units from Main sheet
existing_units_str = str(config_df.loc["Units", 1]) if "Units" in config_df.index else ""
existing_units = set([u.strip() for u in existing_units_str.split(",") if u.strip()])
units_df = pd.read_excel(control_path, sheet_name="Unit")
units_df.columns = [c.strip() for c in units_df.columns]

clinic_units = set()
for _, unit_row in units_df.iterrows():
    unit_code = str(unit_row.get("Units", "")).strip()
    if unit_code not in existing_units:
        continue
    hospital_text = str(unit_row.get("Hospital", "")).strip().lower()
    if hospital_text and "clinic" in hospital_text:
        clinic_units.add(unit_code)

def normalize_report_name(report_name):
    normalized = re.sub(r'[^a-z0-9]+', '_', str(report_name).strip().lower())
    return normalized.strip('_')

def get_units_for_report(report_name):
    normalized = normalize_report_name(report_name)
    if normalized in CLINIC_ONLY_REPORT_NAMES:
        return set(existing_units)
    return set(existing_units) - clinic_units

# === Load Report sheet ===
reports_df = pd.read_excel(control_path, sheet_name="Report")
for col in [status_col, download_col, error_col]:
    if col not in reports_df.columns:
        reports_df[col] = ""
    else:
        if col in (status_col, error_col):
            reports_df[col] = reports_df[col].astype(str)

# === Helper: Convert/fix workbook and ensure first sheet name is "1" ===
def _fix_xlsx_sheet_name(file_path):
    wb = load_workbook(file_path)
    try:
        if wb.sheetnames:
            first_sheet = wb[wb.sheetnames[0]]
            if first_sheet.title != "1":
                first_sheet.title = "1"
        wb.save(file_path)
    finally:
        wb.close()
    return file_path


def convert_and_fix_excel(file_path, retries=5, retry_sleep_s=1.5):
    if not os.path.exists(file_path):
        return None

    ext = os.path.splitext(file_path)[1].lower()

    # xlsx/xlsm: avoid COM entirely; just normalize first sheet name.
    if ext in (".xlsx", ".xlsm"):
        try:
            return _fix_xlsx_sheet_name(file_path)
        except Exception as e:
            # Some ".xlsx" downloads are malformed; attempt Excel re-save to repair.
            last_err = e
            for attempt in range(1, retries + 1):
                excel = None
                wb = None
                pythoncom.CoInitialize()
                try:
                    excel = win32.DispatchEx("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    excel.ScreenUpdating = False
                    excel.EnableEvents = False
                    wb = excel.Workbooks.Open(
                        file_path,
                        UpdateLinks=0,
                        ReadOnly=False,
                        IgnoreReadOnlyRecommended=True,
                        Notify=False,
                    )
                    repaired_path = os.path.splitext(file_path)[0] + ".xlsx"
                    wb.Sheets(1).Name = "1"
                    wb.SaveAs(repaired_path, FileFormat=51)
                    wb.Close(SaveChanges=False)
                    wb = None
                    excel.Quit()
                    excel = None
                    return repaired_path
                except Exception as repair_err:
                    last_err = repair_err
                    err_text = str(repair_err)
                    is_busy = ("0x800ac472" in err_text) or ("-2146777998" in err_text)
                    if attempt < retries and is_busy:
                        time.sleep(retry_sleep_s * attempt)
                        continue
                    break
                finally:
                    try:
                        if wb is not None:
                            wb.Close(SaveChanges=False)
                    except Exception:
                        pass
                    try:
                        if excel is not None:
                            excel.Quit()
                    except Exception:
                        pass
                    pythoncom.CoUninitialize()

            print(f"[WARN] Could not fix xlsx file {file_path}: {last_err}")
            return None

    # xls: convert via Excel COM with retry for transient 0x800ac472 ("Excel busy").
    if ext == ".xls":
        last_err = None
        for attempt in range(1, retries + 1):
            excel = None
            wb = None
            pythoncom.CoInitialize()
            try:
                excel = win32.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.ScreenUpdating = False
                excel.EnableEvents = False

                wb = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=0,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    Notify=False,
                )
                wb.Sheets(1).Name = "1"
                new_file_path = os.path.splitext(file_path)[0] + ".xlsx"
                wb.SaveAs(new_file_path, FileFormat=51)
                wb.Close(SaveChanges=False)
                wb = None
                excel.Quit()
                excel = None

                if file_path != new_file_path and os.path.exists(file_path):
                    os.remove(file_path)
                return new_file_path
            except Exception as e:
                last_err = e
                err_text = str(e)
                is_busy = ("0x800ac472" in err_text) or ("-2146777998" in err_text)
                if attempt < retries and is_busy:
                    time.sleep(retry_sleep_s * attempt)
                    continue
                break
            finally:
                try:
                    if wb is not None:
                        wb.Close(SaveChanges=False)
                except Exception:
                    pass
                try:
                    if excel is not None:
                        excel.Quit()
                except Exception:
                    pass
                pythoncom.CoUninitialize()

        print(f"[WARN] Could not convert/fix file {file_path}: {last_err}")
        return None

    print(f"[WARN] Unsupported file type for conversion: {file_path}")
    return None

# === Helper: Move file to dashboard ===
def move_to_dashboard(src_file, report_name):
    safe_report_name = report_name.replace(" ", "_")
    dest_dir = os.path.join(dashboard_folder, safe_report_name)
    os.makedirs(dest_dir, exist_ok=True)
    dest_file = os.path.join(dest_dir, os.path.basename(src_file))
    shutil.move(src_file, dest_file)
    return dest_file

# === Helper: Delete file safely ===
def delete_file_if_exists(fpath):
    if fpath and os.path.exists(fpath):
        try:
            os.remove(fpath)
        except:
            pass

# === Process all reports marked Download=TRUE ===
download_reports = reports_df[reports_df[download_col].astype(str).str.strip().str.upper() == "TRUE"]
failed_units = set()
total_reports_considered = 0
total_units_considered = 0
total_units_success = 0
total_units_failed = 0

append_run_log(f"[POST] Starting post-process for {len(download_reports)} report(s) marked Download=TRUE.")

for idx, rep in download_reports.iterrows():
    total_reports_considered += 1
    report_name = rep['Report Name']
    report_folder = os.path.join(output_folder, report_name.replace(" ", "_"))
    report_units = sorted(get_units_for_report(report_name))
    append_run_log(f"[POST] Report '{report_name}': expected units={','.join(report_units) if report_units else 'None'}")

    report_failed_units = []
    report_failure_reasons = []

    for unit in report_units:
        total_units_considered += 1
        # Match files: Unit + any FromDate + any ToDate + Report Name
        pattern = os.path.join(report_folder, f"{unit}_*_*_{report_name}*.xls*")
        files = sorted([f for f in glob.glob(pattern) if not f.endswith(".crdownload")])

        if not files:
            report_failed_units.append(unit)
            report_failure_reasons.append(f"{unit}: File missing")
            total_units_failed += 1
            append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': FAILED (File missing)")
            continue

        unit_success_count = 0
        unit_errors = []
        append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': found {len(files)} file(s)")
        for f in files:
            try:
                if not os.path.exists(f):
                    continue
                f = convert_and_fix_excel(f)
                if f is None:
                    unit_errors.append("Conversion failed")
                    append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': conversion failed")
                    continue

                df_dict = pd.read_excel(f, sheet_name=None, engine='openpyxl')
                sheet_df = df_dict["1"]

                if len(sheet_df) < min_rows_required:
                    unit_errors.append(f"Less than {min_rows_required} rows")
                    append_run_log(
                        f"[POST] Report '{report_name}' Unit '{unit}': rejected file "
                        f"(rows={len(sheet_df)} < {min_rows_required})"
                    )
                    continue

                # Success -> move file
                move_to_dashboard(f, report_name)
                unit_success_count += 1
                append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': moved to dashboard")
            except Exception as e:
                unit_errors.append(str(e))
                delete_file_if_exists(f)
                append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': exception={e}")

        if unit_success_count == 0:
            report_failed_units.append(unit)
            reason = unit_errors[-1] if unit_errors else "Unknown processing failure"
            report_failure_reasons.append(f"{unit}: {reason}")
            total_units_failed += 1
            append_run_log(f"[POST] Report '{report_name}' Unit '{unit}': FAILED ({reason})")
            continue
        total_units_success += 1
        append_run_log(
            f"[POST] Report '{report_name}' Unit '{unit}': SUCCESS "
            f"(valid file count={unit_success_count})"
        )

    # Update report row based on units
    if report_failed_units:
        reports_df.at[idx, status_col] = "Not Downloaded"
        reports_df.at[idx, download_col] = True
        reports_df.at[idx, error_col] = "; ".join(report_failure_reasons)
        failed_units.update(report_failed_units)
        append_run_log(
            f"[POST] Report '{report_name}': PARTIAL/FAILED "
            f"(failed units={','.join(sorted(report_failed_units))})"
        )
    else:
        reports_df.at[idx, status_col] = "Downloaded"
        reports_df.at[idx, download_col] = False
        reports_df.at[idx, error_col] = ""
        append_run_log(f"[POST] Report '{report_name}': SUCCESS (all expected units processed)")

# Update Units in Main sheet with remaining failed units
config_df.loc["Units", 1] = ",".join(sorted(failed_units)) if failed_units else ""

# Save sheets
with pd.ExcelWriter(control_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    reports_df.to_excel(writer, sheet_name="Report", index=False)
    config_df.to_excel(writer, sheet_name="Main", header=False)

print("[OK] Post-processing completed. Units, Status, and Failure Reason updated correctly.")
append_run_log(
    "[POST] Completed post-process: "
    f"reports={total_reports_considered}, units={total_units_considered}, "
    f"unit_success={total_units_success}, unit_failed={total_units_failed}, "
    f"remaining_failed_units={','.join(sorted(failed_units)) if failed_units else 'None'}"
)


